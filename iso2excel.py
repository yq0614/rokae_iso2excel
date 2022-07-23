# _*_ encoding：utf-8 _*_
import json
import re
# import sys
import pdfplumber
from pdfminer.pdftypes import *
import openpyxl as op

with open('table.json', 'r', encoding='utf8') as fp:
    data = json.load(fp)
    list_key = []
    for i in data.keys():
        list_key.append(i)

with open('name.json', 'r', encoding='utf8') as fp:
    name = json.load(fp)
    list_name = []
    for i in name.keys():
        list_name.append(i)


def gettable():
    pdf = pdfplumber.open(sys.argv[1])
    content = ""
    for page in pdf.pages:
        content += page.extract_text()
    original = content.replace("\n", " ")
    original = original.replace("%", " ")
    result = []
    for n in range(0, len(data), 2):
        compile_str = re.compile(data[list_key[n]] + ".*?" + data[list_key[n+1]], re.M | re.I)
        match = compile_str.findall(original)
        if match == [ ]:
            result.append('\n')
        else:
            str_match = match[0].replace(data[list_key[n]], "").replace(data[list_key[n + 1]], "")
            result.append(str_match)
    return result


def write2excel(need_data):
    book = op.Workbook()
    sheet = book.active
    sheet.title = "iso数据提取"
    for rs, rl in [[1, 5], [6, 10], [11, 13]]:
        sheet.merge_cells(start_row=rs, end_row=rl, start_column=1, end_column=1)
    colA = sheet.column_dimensions['A']
    colA.width = 25
    column2 = ['P1', 'P2', 'P3', 'P4', 'P5']*2 + ['P1', 'P2', 'P4'] + ['']*13
    for row in range(1, len(need_data)+1):
        sheet.cell(row, 3, need_data[row-1])
        sheet.cell(row, 2, column2[row - 1])
        if row > 13:
            sheet.cell(row, 1, name[list_name[row - 1]])
    for k in (1, 6, 11):
        sheet.cell(k, 1, name[list_name[k - 1]])
    book.save(sys.argv[1][:-4] + '.xlsx')


def get_need_data(ori):
    tmp = []
    ls = [25, 9, 4, 19, 3, 4, 2, 4]
    for j in ori:
        tmp.append(j.split(" ")[1:-1])
    for i in range(0, len(tmp)):
        if tmp[i] == []:
            tmp[i] = ['none']*ls[i]
    need_data = tmp[0][1:25:5] + tmp[0][2:25:5] + tmp[1][1:9:3] + tmp[2][0:2] + tmp[3][-2:] + tmp[4][:2] + tmp[5][:2] + tmp[6] + tmp[7][1:]
    return need_data


write2excel(get_need_data(gettable()))